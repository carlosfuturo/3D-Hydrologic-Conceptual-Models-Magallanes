"""
Script 12 – Dataset Estratigráfico con Unidades Hidrogeológicas (UH)
======================================================================
Construye el dataset de estratigrafía de pozos en formato de referencia:
  Pozo | Prof_Desde_m | Prof_Hasta_m | Estratigrafia | UH | UH_Descripcion |
  Este_WGS84_H19S | Norte_WGS84_H19S

Fuentes:
  • ND-1202-1299  Morro Chico (Penitente) – extracción previa de PDF
  • ND-1203-765   EA. Arizona (Tierra del Fuego) – extracción de PDF
  • ND-1203-723   Cabaña Sur ZG-3 (Bloque Arenal) – interpretación 2 unidades
  • ND-1203-724   Lautaro Sur 7  (Bloque Arenal) – interpretación 2 unidades
  • ND-1203-725   Cabaña ZG-3   (Bloque Arenal) – interpretación 2 unidades

UH (Unidades Hidrogeológicas) clasificadas:
  UH-0  Suelo orgánico / cobertura vegetal         → No acuífero
  UH-1  Depósitos fluvioglaciares (gravas/arenas)  → Acuífero libre
  UH-2  Arcillas y limos                           → Acuítardo / Acuicludo
  UH-3  Mezcla arena-arcilla-grava-limo            → Acuífero baja permeabilidad
  UH-4  Areniscas / sedimentitas (no confinadas)   → Acuífero sedimentario
  UH-5  Areniscas confinadas / Fm. Palomares       → Acuífero confinado artesiano
  UH-6  Arcilla compacta / roca basal              → Basamento impermeable

Salidas:
  ├── dataset_estratigrafico_UH.xlsx  (coloreado por UH)
  ├── dataset_estratigrafico_UH.csv
  └── Figuras_Pozos_Acuifero/Fig12_Columnas_Estratigraficas.png
"""

import os
import sys
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 0. RUTAS
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(BASE_DIR, "dataset_estratigrafico_UH.xlsx")
OUT_CSV  = os.path.join(BASE_DIR, "dataset_estratigrafico_UH.csv")
FIG_DIR  = os.path.join(BASE_DIR, "Figuras_Pozos_Acuifero")
OUT_FIG  = os.path.join(FIG_DIR, "Fig12_Columnas_Estratigraficas.png")
os.makedirs(FIG_DIR, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# 1. DATOS ESTRATIGRÁFICOS (hardcoded de PDFs)
# ─────────────────────────────────────────────────────────────────────────────
# Cada entrada: (Pozo, Desde, Hasta, Litología, E_wgs84, N_wgs84, fuente)
# Nota: Cuenca Penitente = WGS84 Huso 19S; Tierra del Fuego = idem

RAW_DATA = [

    # ── ND-1202-1299  Estancia Morro Chico (Penitente) ───────────────────────
    # Fuente: Memoria Explicativa ND-1202-1299, extracción previa
    # Coord WGS84 H19S: E=331957  N=4230484   Prof=43 m  NE=7.4 m  Q=1.2 L/s
    ("ND-1202-1299",  0.0,  0.3,  "Capa vegetal",                331957.0, 4230484.0),
    ("ND-1202-1299",  0.3,  3.0,  "Arcilla amarilla",             331957.0, 4230484.0),
    ("ND-1202-1299",  3.0, 12.0,  "Arena y grava",                331957.0, 4230484.0),
    ("ND-1202-1299", 12.0, 42.0,  "Grava y piedra (acuífero)",    331957.0, 4230484.0),
    ("ND-1202-1299", 42.0, 43.0,  "Grava y arcilla gris",         331957.0, 4230484.0),

    # ── ND-1203-765  EA. Arizona – China Creek (Tierra del Fuego) ────────────
    # Fuente: "Informe pozo Arizona (1).pdf"  – Columna litográfica del pozo
    # Coord WGS84 H19S: E=500362  N=4100774   Prof=48 m  NE=6.5 m  Q=1.5 L/s
    ("ND-1203-765",  0.0,  1.5,  "Perfil orgánico",                500362.0, 4100774.0),
    ("ND-1203-765",  1.5,  4.0,  "Arcilla café",                   500362.0, 4100774.0),
    ("ND-1203-765",  4.0,  6.0,  "Arcilla café - gravilla",        500362.0, 4100774.0),
    ("ND-1203-765",  6.0,  9.0,  "Arcilla café - arena gruesa",    500362.0, 4100774.0),
    ("ND-1203-765",  9.0, 11.0,  "Arenisca café",                  500362.0, 4100774.0),
    ("ND-1203-765", 11.0, 13.0,  "Arenisca gris",                  500362.0, 4100774.0),
    ("ND-1203-765", 13.0, 13.5,  "Arena fina",                     500362.0, 4100774.0),
    ("ND-1203-765", 13.5, 22.0,  "Arenisca gris",                  500362.0, 4100774.0),
    ("ND-1203-765", 22.0, 25.0,  "Arcilla gris - arena fina",      500362.0, 4100774.0),
    ("ND-1203-765", 25.0, 32.0,  "Arcilla gris",                   500362.0, 4100774.0),
    ("ND-1203-765", 32.0, 36.0,  "Arena gruesa",                   500362.0, 4100774.0),
    ("ND-1203-765", 36.0, 38.0,  "Limo - arena",                   500362.0, 4100774.0),
    ("ND-1203-765", 38.0, 42.0,  "Arena - gravilla",               500362.0, 4100774.0),
    ("ND-1203-765", 42.0, 46.0,  "Arenisca",                       500362.0, 4100774.0),
    ("ND-1203-765", 46.0, 48.0,  "Arcilla café",                   500362.0, 4100774.0),

    # ── ND-1203-723  Cabaña Sur ZG-3 (Bloque Arenal, Tierra del Fuego) ───────
    # Fuente: "Estudio Aguas Cabaña Sur.pdf" (IPROMSA, 2018)
    # Estratigrafía: 2 unidades de Figura 5 (label de imagen)
    #   "Base de los Depósitos Fluvioglaciares / Límite Superior Fm. Palomares"
    # Screen: 114.81–138.64 m → contacto a ~115 m
    # Coord WGS84 H19S: E=450184  N=4144485  Prof=140.5 m  Q=0.87 L/s  SURGENTE
    ("ND-1203-723",   0.0, 115.0, "Depósitos Fluvioglaciares (gravas, arenas, limos glaciales)", 450184.0, 4144485.0),
    ("ND-1203-723", 115.0, 140.5, "Formación Palomares (areniscas confinadas – nivel artesiano)", 450184.0, 4144485.0),

    # ── ND-1203-724  Lautaro Sur 7 (Bloque Arenal, Tierra del Fuego) ─────────
    # Fuente: "Estudio Pozo Lautaro (La Kutu).pdf" (IPROMSA, 2018)
    # Screen: 38.00–78.54 m → contacto a ~38 m
    # Coord WGS84 H19S: E=458714  N=4142414  Prof=80.5 m  Q=8.43 L/s  SURGENTE
    ("ND-1203-724",  0.0, 38.0,  "Depósitos Fluvioglaciares (gravas, arenas, limos glaciales)", 458714.0, 4142414.0),
    ("ND-1203-724", 38.0, 80.5,  "Formación Palomares (areniscas confinadas – nivel artesiano)", 458714.0, 4142414.0),

    # ── ND-1203-725  Cabaña ZG-3 (Bloque Arenal, Tierra del Fuego) ───────────
    # Fuente: "Estudio Cabaña ZG-3.pdf" (IPROMSA, 2018)
    # Screen: 160.90–182.69 m → contacto a ~161 m
    # Coord WGS84 H19S: E=449672  N=4148398  Prof=186.5 m  Q=0.55 L/s  SURGENTE
    ("ND-1203-725",   0.0, 161.0, "Depósitos Fluvioglaciares (gravas, arenas, limos glaciales)", 449672.0, 4148398.0),
    ("ND-1203-725", 161.0, 186.5, "Formación Palomares (areniscas confinadas – nivel artesiano)", 449672.0, 4148398.0),
]

# ─────────────────────────────────────────────────────────────────────────────
# 2. CLASIFICACIÓN DE UH
# ─────────────────────────────────────────────────────────────────────────────

UH_DEFS = {
    "UH-0": "Suelo orgánico / Cobertura vegetal",
    "UH-1": "Depósitos fluvioglaciares permeables – Acuífero libre",
    "UH-2": "Arcillas y limos – Acuítardo / Acuicludo",
    "UH-3": "Mezcla arena-arcilla / grava-limo – Acuífero baja permeabilidad",
    "UH-4": "Areniscas y sedimentitas – Acuífero sedimentario",
    "UH-5": "Formación Palomares (areniscas confinadas) – Acuífero artesiano",
    "UH-6": "Roca compacta / Arcilla dura – Basamento impermeable",
}

UH_COLORS_HEX = {
    "UH-0": "F5CBA7",   # beige claro  – suelo vegetal
    "UH-1": "F9E79F",   # amarillo – acuífero libre (gravas)
    "UH-2": "D2B4DE",   # lila – acuítardo arcilla
    "UH-3": "FAD7A0",   # naranja claro – mezcla arena-arcilla
    "UH-4": "AED6F1",   # azul claro – areniscas
    "UH-5": "85C1E9",   # azul – Fm. Palomares confinado
    "UH-6": "AEB6BF",   # gris – roca/arcilla dura
}

import re

def clasifica_UH(litologia: str) -> str:
    """Asigna una UH a una descripción litológica."""
    s = litologia.lower()

    # UH-0: suelo vegetal / perfil orgánico
    if re.search(r"vegetal|org[aá]nic|humus|capa vegetal|perfil org", s):
        return "UH-0"

    # UH-5: Formación Palomares (artesiano) — prioridad alta
    if "palomares" in s or "confinada" in s or "artesian" in s:
        return "UH-5"

    # UH-1: depósitos fluvioglaciares reconocidos por nombre — prioridad alta
    # (puede contener "limos" como limos glaciales pero sigue siendo UH-1)
    if re.search(r"fluvioglaciar", s):
        return "UH-1"

    # UH-2: arcilla pura / limo puro (sin gravas ni arenas)
    if re.search(r"\barcilla\b|\blimo\b|\blimosa\b|\bgreda\b", s) \
       and not re.search(r"grava|arena|ripio|bol[oó]n|gravilla|arenisca", s):
        return "UH-2"

    # UH-1: gravas, ripios, bolones, piedras sin finos
    if re.search(r"\bgrava\b|\bripio\b|\bbol[oó]n\b|\bpiedra\b|\bgravilla\b", s) \
       and not re.search(r"\barcilla\b|\blimo\b", s):
        return "UH-1"

    # UH-1: arena / arena gruesa sin finos ni arenisca
    if re.search(r"\barena\b|\barena gruesa\b|\barena fina\b|\barena - gravilla\b", s) \
       and not re.search(r"\barcilla\b|\blimo\b|\barenisca\b", s):
        return "UH-1"

    # UH-4: areniscas (sedimentitas no confinadas)
    if re.search(r"\barenisca\b", s):
        return "UH-4"

    # UH-3: mezclas arena-arcilla, grava-limo, arcilla-arena, etc.
    if (re.search(r"\barena\b|\bgrava\b|\bgravilla\b", s)
            and re.search(r"\barcilla\b|\blimo\b", s)):
        return "UH-3"

    # UH-3 fallback para limo-arena o combinaciones
    if re.search(r"limo.*arena|arena.*limo", s):
        return "UH-3"

    # UH-6: roca dura, basamento
    if re.search(r"\broca\b|basamento|cemented", s):
        return "UH-6"

    return "UH-3"  # por defecto: mezcla / indeterminado

# ─────────────────────────────────────────────────────────────────────────────
# 3. CONSTRUIR DATAFRAME
# ─────────────────────────────────────────────────────────────────────────────

records = []
for (pozo, desde, hasta, litol, este, norte) in RAW_DATA:
    uh = clasifica_UH(litol)
    records.append({
        "Pozo":              pozo,
        "Prof_Desde_m":      desde,
        "Prof_Hasta_m":      hasta,
        "Estratigrafia":     litol,
        "UH":                uh,
        "UH_Descripcion":    UH_DEFS[uh],
        "Este UTM WGS84 H19S":   este,
        "Norte UTM WGS84 H19S":  norte,
    })

df = pd.DataFrame(records)

# ─────────────────────────────────────────────────────────────────────────────
# 4. EXPORTAR CSV
# ─────────────────────────────────────────────────────────────────────────────
df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
print(f"[OK] CSV guardado: {OUT_CSV}")
print(f"     {len(df)} registros | {df['Pozo'].nunique()} pozos")

# ─────────────────────────────────────────────────────────────────────────────
# 5. EXPORTAR XLSX CON COLORES POR UH
# ─────────────────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dataset_UH"

# ── Encabezados ──────────────────────────────────────────────────────────────
headers = [
    "Pozo", "Prof_Desde_m", "Prof_Hasta_m", "Estratigrafia",
    "UH", "UH_Descripcion", "Este UTM WGS84 H19S", "Norte UTM WGS84 H19S"
]
header_fill = PatternFill("solid", fgColor="2C3E50")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="000000")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border_all

ws.row_dimensions[1].height = 30

# ── Ancho de columnas ─────────────────────────────────────────────────────────
col_widths = {1: 20, 2: 13, 3: 13, 4: 48, 5: 8, 6: 52, 7: 16, 8: 16}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# ── Datos con color por UH ────────────────────────────────────────────────────
data_align_center = Alignment(horizontal="center", vertical="center")
data_align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

for row_idx, row in df.iterrows():
    r = row_idx + 2   # fila Excel (1=header)
    uh_code = row["UH"]
    fill_color = UH_COLORS_HEX.get(uh_code, "FFFFFF")
    row_fill = PatternFill("solid", fgColor=fill_color)

    values = [
        row["Pozo"],
        row["Prof_Desde_m"],
        row["Prof_Hasta_m"],
        row["Estratigrafia"],
        row["UH"],
        row["UH_Descripcion"],
        row["Este UTM WGS84 H19S"],
        row["Norte UTM WGS84 H19S"],
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.fill = row_fill
        cell.border = border_all
        # Centro para numéricos / izquierda para texto
        if col_idx in (2, 3, 5, 7, 8):
            cell.alignment = data_align_center
        else:
            cell.alignment = data_align_left
        # Negrilla para código de pozo
        if col_idx == 1:
            cell.font = Font(bold=True)
        # Negrita leve para UH
        if col_idx == 5:
            cell.font = Font(bold=True)

# ── Fila de leyenda ───────────────────────────────────────────────────────────
last_data_row = len(df) + 2 + 1  # una fila en blanco

legend_header = ws.cell(row=last_data_row, column=1, value="LEYENDA UH")
legend_header.font = Font(bold=True, size=11)
legend_header.alignment = data_align_center

for i, (code, desc) in enumerate(UH_DEFS.items()):
    lr = last_data_row + 1 + i
    c1 = ws.cell(row=lr, column=1, value=code)
    c2 = ws.cell(row=lr, column=2, value=desc)
    fill = PatternFill("solid", fgColor=UH_COLORS_HEX[code])
    c1.fill = fill
    c2.fill = fill
    c1.font = Font(bold=True)
    c1.alignment = data_align_center
    c2.alignment = data_align_left
    ws.merge_cells(start_row=lr, start_column=2, end_row=lr, end_column=6)

# ── Guardar ───────────────────────────────────────────────────────────────────
wb.save(OUT_XLSX)
print(f"[OK] XLSX guardado: {OUT_XLSX}")

# ─────────────────────────────────────────────────────────────────────────────
# 6. FIGURA: COLUMNAS ESTRATIGRÁFICAS MULTI-POZO
# ─────────────────────────────────────────────────────────────────────────────

UH_COLORS_MPL = {
    "UH-0": "#F5CBA7",
    "UH-1": "#F9E79F",
    "UH-2": "#D2B4DE",
    "UH-3": "#FAD7A0",
    "UH-4": "#AED6F1",
    "UH-5": "#85C1E9",
    "UH-6": "#AEB6BF",
}

pozos_order = df["Pozo"].unique().tolist()
n_pozos = len(pozos_order)

# Calcular profundidad máxima para escala
max_prof = df["Prof_Hasta_m"].max()

# Tamaño: un panel por pozo + espacio leyenda
fig_w = 3.5 * n_pozos + 3.5
fig_h = max(10, max_prof * 0.08 + 4)
fig, axes = plt.subplots(1, n_pozos, figsize=(fig_w, fig_h),
                         gridspec_kw={"wspace": 0.05})
fig.patch.set_facecolor("#F0F3F4")

if n_pozos == 1:
    axes = [axes]

# ── Título ────────────────────────────────────────────────────────────────────
fig.suptitle(
    "Columnas Estratigráficas y Unidades Hidrogeológicas\n"
    "Cuencas Penitente y El Oro – WGS84 Huso 19S",
    fontsize=14, fontweight="bold", y=0.98, color="#1B2631"
)

for ax_idx, pozo in enumerate(pozos_order):
    ax = axes[ax_idx]
    sub = df[df["Pozo"] == pozo].sort_values("Prof_Desde_m")

    # Prof máxima para este pozo
    pmax = sub["Prof_Hasta_m"].max()

    ax.set_xlim(0, 1)
    ax.set_ylim(pmax + 1, -1)  # eje Y invertido
    ax.set_facecolor("#FDFEFE")

    # Dibujar bloques litológicos
    for _, fila in sub.iterrows():
        d = fila["Prof_Desde_m"]
        h = fila["Prof_Hasta_m"]
        uh = fila["UH"]
        color = UH_COLORS_MPL.get(uh, "#FFFFFF")
        grosor = h - d
        rect = mpatches.FancyBboxPatch(
            (0.05, d), 0.90, grosor,
            boxstyle="square,pad=0",
            linewidth=0.6, edgecolor="#555555",
            facecolor=color, zorder=2
        )
        ax.add_patch(rect)

        # Texto dentro del bloque
        mid = (d + h) / 2
        lit_txt = fila["Estratigrafia"]
        # Acortar texto largo
        if len(lit_txt) > 30:
            lit_txt = lit_txt[:28] + "…"
        fontsize = 6.5 if grosor < 2 else 7.5
        ax.text(0.50, mid, lit_txt, ha="center", va="center",
                fontsize=fontsize, color="#1B2631", zorder=3,
                wrap=True, style="italic")
        # Etiqueta UH a la derecha
        ax.text(0.96, mid, uh, ha="right", va="center",
                fontsize=6.5, fontweight="bold", color="#2C3E50", zorder=3)
        # Líneas de profundidad
        ax.axhline(d, color="#AAAAAA", linewidth=0.4, zorder=1)
        ax.axhline(h, color="#AAAAAA", linewidth=0.4, zorder=1)

    # eje Y: profundidades
    yticks = list(range(0, int(pmax) + 5, 5))
    ax.set_yticks(yticks)
    if ax_idx == 0:
        ax.set_yticklabels([f"{y} m" for y in yticks], fontsize=8)
    else:
        ax.set_yticklabels([])

    # eje X oculto
    ax.set_xticks([])

    # título del panel = código pozo + coordenadas
    coord_row = sub.iloc[0]
    e_str = f"E={coord_row['Este UTM WGS84 H19S']:,.0f}"
    n_str = f"N={coord_row['Norte UTM WGS84 H19S']:,.0f}"
    ax.set_title(f"{pozo}\n{e_str}\n{n_str}",
                 fontsize=8, fontweight="bold", color="#1B2631",
                 pad=6, linespacing=1.3)

    # caja del panel
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color("#888888")
        spine.set_linewidth(0.8)

# ── Leyenda ───────────────────────────────────────────────────────────────────
legend_handles = []
for code, color in UH_COLORS_MPL.items():
    label = f"{code}: {UH_DEFS[code][:45]}…" if len(UH_DEFS[code]) > 45 else f"{code}: {UH_DEFS[code]}"
    patch = mpatches.Patch(facecolor=color, edgecolor="#555555", linewidth=0.8, label=label)
    legend_handles.append(patch)

fig.legend(
    handles=legend_handles,
    loc="lower center",
    ncol=2,
    fontsize=8,
    frameon=True,
    framealpha=0.9,
    edgecolor="#888888",
    title="Unidades Hidrogeológicas (UH)",
    title_fontsize=9,
    bbox_to_anchor=(0.5, 0.01),
)

plt.tight_layout(rect=[0, 0.13, 1, 0.96])
fig.savefig(OUT_FIG, dpi=180, bbox_inches="tight",
            facecolor=fig.get_facecolor())
plt.close(fig)
print(f"[OK] Figura guardada: {OUT_FIG}")

# ─────────────────────────────────────────────────────────────────────────────
# 7. RESUMEN EN CONSOLA
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("DATASET ESTRATIGRÁFICO – RESUMEN")
print("=" * 70)
print(f"{'Pozo':<20} {'N° capas':>8} {'Prof_max':>9} {'Este':>10} {'Norte':>12}")
print("-" * 70)
for pozo, grp in df.groupby("Pozo", sort=False):
    print(f"{pozo:<20} {len(grp):>8} {grp['Prof_Hasta_m'].max():>8.1f}m "
          f"{grp['Este UTM WGS84 H19S'].iloc[0]:>10.0f} {grp['Norte UTM WGS84 H19S'].iloc[0]:>12.0f}")

print("\nDistribución UH:")
uh_cnt = df["UH"].value_counts().sort_index()
for uh, cnt in uh_cnt.items():
    print(f"  {uh}: {cnt:>3} capa(s) – {UH_DEFS[uh]}")

print(f"\nArchivos generados:")
print(f"  {OUT_CSV}")
print(f"  {OUT_XLSX}")
print(f"  {OUT_FIG}")
